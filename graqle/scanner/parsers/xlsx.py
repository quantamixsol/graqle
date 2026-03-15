"""XLSX spreadsheet parser — requires ``openpyxl``.

Treats each worksheet as a section.  Rows are stored as table data.
"""

# ── graqle:intelligence ──
# module: graqle.scanner.parsers.xlsx
# risk: LOW (impact radius: 1 modules)
# consumers: test_xlsx
# dependencies: __future__, pathlib, typing, base
# constraints: none
# ── /graqle:intelligence ──

from __future__ import annotations

from pathlib import Path
from typing import Any

from graqle.scanner.parsers.base import (
    BaseDocParser,
    ParsedDocument,
    ParsedSection,
)


class XLSXParser(BaseDocParser):
    """Parse ``.xlsx`` files using ``openpyxl``."""

    def parse(self, path: Path) -> ParsedDocument:
        """Parse an XLSX file into a :class:`ParsedDocument`."""
        import openpyxl

        path = Path(path)
        if not path.is_file():
            raise FileNotFoundError(f"XLSX file not found: {path}")

        metadata: dict[str, Any] = {"source": str(path)}
        parse_errors: list[str] = []

        try:
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError(f"Failed to parse XLSX: {exc}") from exc

        metadata["sheet_count"] = len(wb.sheetnames)

        sections: list[ParsedSection] = []
        full_text_parts: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_data: list[list[str]] = []

            try:
                for row in ws.iter_rows(values_only=True):
                    row_vals = [str(c) if c is not None else "" for c in row]
                    # Skip entirely empty rows
                    if any(v.strip() for v in row_vals):
                        rows_data.append(row_vals)
            except Exception as exc:
                parse_errors.append(f"Sheet '{sheet_name}': {exc}")
                continue

            if not rows_data:
                continue

            # First row as headers
            headers = rows_data[0]
            data_rows = rows_data[1:] if len(rows_data) > 1 else []

            table = {"headers": headers, "rows": data_rows}

            # Build text representation
            text_lines = [" | ".join(headers)]
            for row in data_rows[:50]:  # Limit to first 50 rows for text
                text_lines.append(" | ".join(row))

            content = "\n".join(text_lines)
            full_text_parts.append(f"{sheet_name}\n{content}")

            sections.append(
                ParsedSection(
                    title=sheet_name,
                    content=content,
                    level=1,
                    section_type="sheet",
                    tables=[table],
                    code_blocks=[],
                    links=[],
                )
            )

        try:
            wb.close()
        except Exception:
            pass

        title = path.stem
        full_text = "\n\n".join(full_text_parts)

        return ParsedDocument(
            path=path,
            title=title,
            format="xlsx",
            sections=sections,
            full_text=full_text,
            metadata=metadata,
            parse_errors=parse_errors,
        )

    def is_available(self) -> bool:
        try:
            import openpyxl  # noqa: F401
            return True
        except ImportError:
            return False

    def missing_dependency_message(self) -> str:
        return "Install openpyxl: pip install graqle[docs]"

    @property
    def supported_extensions(self) -> list[str]:
        return [".xlsx"]
