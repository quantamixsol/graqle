"""Tests for graqle.scanner.parsers.xlsx — XLSX parser."""

# ── graqle:intelligence ──
# module: tests.test_scanner.test_parsers.test_xlsx
# risk: LOW (impact radius: 0 modules)
# dependencies: __future__, pathlib, pytest, xlsx
# constraints: none
# ── /graqle:intelligence ──

from __future__ import annotations

from pathlib import Path

import pytest

openpyxl = pytest.importorskip("openpyxl", reason="openpyxl not installed")

from graqle.scanner.parsers.xlsx import XLSXParser


@pytest.fixture
def parser() -> XLSXParser:
    return XLSXParser()


@pytest.fixture
def sample_xlsx(tmp_path: Path) -> Path:
    """Create a minimal XLSX file."""
    from openpyxl import Workbook

    p = tmp_path / "test.xlsx"
    wb = Workbook()

    # Sheet 1
    ws1 = wb.active
    ws1.title = "Services"
    ws1.append(["Name", "Type", "Status"])
    ws1.append(["auth_service", "MODULE", "active"])
    ws1.append(["db_pool", "SERVICE", "active"])
    ws1.append(["api_handler", "FUNCTION", "active"])

    # Sheet 2
    ws2 = wb.create_sheet("Config")
    ws2.append(["Key", "Value"])
    ws2.append(["DATABASE_URL", "postgresql://localhost/db"])
    ws2.append(["JWT_SECRET", "super-secret"])

    wb.save(str(p))
    return p


@pytest.fixture
def empty_xlsx(tmp_path: Path) -> Path:
    """Create an empty XLSX file."""
    from openpyxl import Workbook

    p = tmp_path / "empty.xlsx"
    wb = Workbook()
    wb.save(str(p))
    return p


class TestXLSXParser:
    def test_is_available(self, parser: XLSXParser) -> None:
        assert parser.is_available() is True

    def test_supported_extensions(self, parser: XLSXParser) -> None:
        assert ".xlsx" in parser.supported_extensions

    def test_parse_basic(self, parser: XLSXParser, sample_xlsx: Path) -> None:
        doc = parser.parse(sample_xlsx)
        assert doc.format == "xlsx"
        assert doc.path == sample_xlsx
        assert len(doc.sections) >= 2  # Two sheets

    def test_parse_metadata(self, parser: XLSXParser, sample_xlsx: Path) -> None:
        doc = parser.parse(sample_xlsx)
        assert doc.metadata.get("sheet_count") >= 2

    def test_sheets_as_sections(self, parser: XLSXParser, sample_xlsx: Path) -> None:
        doc = parser.parse(sample_xlsx)
        titles = [s.title for s in doc.sections]
        assert "Services" in titles
        assert "Config" in titles

    def test_section_type_is_sheet(self, parser: XLSXParser, sample_xlsx: Path) -> None:
        doc = parser.parse(sample_xlsx)
        for section in doc.sections:
            assert section.section_type == "sheet"

    def test_table_extraction(self, parser: XLSXParser, sample_xlsx: Path) -> None:
        doc = parser.parse(sample_xlsx)
        services_section = [s for s in doc.sections if s.title == "Services"][0]
        assert len(services_section.tables) >= 1
        table = services_section.tables[0]
        assert "Name" in table["headers"]
        assert len(table["rows"]) >= 3

    def test_full_text_content(self, parser: XLSXParser, sample_xlsx: Path) -> None:
        doc = parser.parse(sample_xlsx)
        assert "auth_service" in doc.full_text

    def test_empty_xlsx(self, parser: XLSXParser, empty_xlsx: Path) -> None:
        doc = parser.parse(empty_xlsx)
        assert doc.format == "xlsx"
        # Empty sheets may still produce sections with just headers

    def test_file_not_found(self, parser: XLSXParser, tmp_path: Path) -> None:
        with pytest.raises(FileNotFoundError):
            parser.parse(tmp_path / "nonexistent.xlsx")
